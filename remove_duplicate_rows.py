from openpyxl import load_workbook, Workbook

def remove_duplicate_rows(input_file, output_file):
    """
    Removes duplicate rows from an Excel file and saves the cleaned dataset.

    :param input_file: Path to the input Excel file.
    :param output_file: Path to save the cleaned Excel file.
    """
    # Load the workbook and select the active worksheet
    wb = load_workbook(input_file)
    ws = wb.active

    # Create a new workbook for the cleaned data
    cleaned_wb = Workbook()
    cleaned_ws = cleaned_wb.active

    # Keep track of seen rows
    seen_rows = set()

    # Iterate through rows and add unique rows to the cleaned dataset
    for row in ws.iter_rows(values_only=True):
        if row not in seen_rows:
            cleaned_ws.append(row)
            seen_rows.add(row)

    # Save the cleaned workbook
    cleaned_wb.save(output_file)
    print(f"Cleaned dataset with duplicates removed saved to {output_file}")

# Example usage
# input_file = "input_dataset.xlsx"
# output_file = "cleaned_dataset.xlsx"

# remove_duplicate_rows(input_file, output_file)

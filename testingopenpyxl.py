from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

current_work_book = "test_date_format.xlsx"

wb = load_workbook(current_work_book)

# wb = Workbook()

ws = wb.active

data = {
    "Joe": {
        "math": 65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Bill": {
        "math": 55,
        "science": 72,
        "english": 87,
        "gym": 95
    },
    "Tim": {
        "math": 100,
        "science": 45,
        "english": 75,
        "gym": 56
    },
    "Jane": {
        "math": 100,
        "science": 100,
        "english": 100,
        "gym": 60
    }
}


# ws = wb["Date Sheet"]  #specific worksheet in the loaded workbook

ws["A1"].value = "Openpyxl works"

# wb.create_sheet("New Sheet")

def custom_filter (list, item):
    if item in list:
        return True
    else: 
        return False 

sheet_exists = custom_filter(wb.sheetnames, "New Sheet")

if not sheet_exists:
    wb.create_sheet("New Sheet")
    

# ws.title = "Mod Date Sheet"

# ws.append(["SN", "NAME", "ADDRESS", "PHONE"])

# ws.merge_cells("A1:D4") # to merge cells in a given range
# ws.unmerge_cells("A1:D1") # to unmerge cells in a given range
# ws.insert_rows(7) # to insert a row in position 7
# ws.delete_rows(7) # to delete a row in position 7
# ws.insert_cols(7) # to insert a column in position 7
# ws.move_range("C1:D11", rows=2, cols=-2) # to move range of cells C1 through D11 by 2 rows down and 2 cols left

for row in range(1, 11):
    for col in range(1, 5):
        # char = chr(65 + col) # or use
        char = get_column_letter(col)
        # print(ws[char + str(row)])       
        # print(ws[char + str(row)].value)       
        ws[char + str(row)] = char + str(row)  

# ws = wb["New Sheet"]
ws = wb["Grades"]

# ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
# headings = ['Name'] + data.keys()

ws.append(headings)

for person in data: 
    grades = data[person].values()
    ws.append([person] + list(grades))     

try:
    wb.save(current_work_book)
except:
    print("file is open, hence cannot perform save operation")

print(wb.sheetnames)

print(ws)

print(len(wb.sheetnames))

print(ws["A1"].value)

print(ws["F2"].value)


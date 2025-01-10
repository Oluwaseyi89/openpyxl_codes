from openpyxl import load_workbook, Workbook

def filter_out_subset(superset_file, subset_file, output_file, key_columns):
    """
    Filters out rows in the superset dataset that match rows in the subset dataset based on key columns.

    :param superset_file: Path to the superset Excel file.
    :param subset_file: Path to the subset Excel file.
    :param output_file: Path to save the resulting filtered dataset.
    :param key_columns: List of column names (headers) to use for matching rows.
    """
    # Load superset and subset workbooks
    superset_wb = load_workbook(superset_file)
    subset_wb = load_workbook(subset_file)

    superset_ws = superset_wb.active
    subset_ws = subset_wb.active

    # Extract headers and ensure key columns exist in both datasets
    superset_header = [cell.value for cell in superset_ws[1]]
    subset_header = [cell.value for cell in subset_ws[1]]

    for key in key_columns:
        if key not in superset_header or key not in subset_header:
            # raise ValueError(f"Key column '{key}' not found in both datasets.")
            pass

    # Get indices of key columns in both datasets
    superset_key_indices = [superset_header.index(key) for key in key_columns]
    subset_key_indices = [subset_header.index(key) for key in key_columns]

    # Build a set of key tuples from the subset dataset
    # subset_keys = set(
    #     tuple(row[idx] for idx in subset_key_indices)
    #     for row in subset_ws.iter_rows(min_row=2, values_only=True)
    #     if row  # Ensure the row is not None
    # )
    
    subset_keys = set(
    tuple(row[idx] for idx in subset_key_indices)
    for row in subset_ws.iter_rows(min_row=2, values_only=True)
    if row and all(row[idx] is not None for idx in subset_key_indices)  # Ensure the row and key columns are not None
)

    print(subset_keys)

    # Create a new workbook for the filtered data
    filtered_wb = Workbook()
    filtered_ws = filtered_wb.active
    filtered_ws.append(superset_header)  # Write the header

    # Add rows from the superset that are not in the subset
    for row in superset_ws.iter_rows(min_row=2, values_only=True):
        if row:  # Skip empty rows
            superset_key = tuple(row[idx] for idx in superset_key_indices)
            if superset_key not in subset_keys:
                filtered_ws.append(row)

    # Save the filtered workbook
    filtered_wb.save(output_file)
    print(f"Filtered dataset saved to {output_file}")

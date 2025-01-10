from openpyxl import load_workbook

def fill_blank_spaces(file_path, output_file):
    """
    Fills blank spaces in an Excel file with zeros for numeric types
    and dashes for other types.

    :param file_path: Path to the input Excel file.
    :param output_file: Path to save the modified Excel file.
    """
    # Load the workbook and select the active worksheet
    wb = load_workbook(file_path)
    ws = wb.active

    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell is blank (None or empty string)
            if cell.value in (None, ""):
                # Fill with 0 if the column is numeric, otherwise fill with "-"
                if cell.column_letter and ws[cell.column_letter + "1"].value:
                    # Replace blanks with appropriate value based on cell data type
                    if isinstance(cell.value, (int, float)):
                        cell.value = 0
                    else:
                        cell.value = '-'

    # Save the modified workbook
    wb.save(output_file)
    print(f"Updated file saved to {output_file}")


# Example usage
# input_file = "input_dataset.xlsx"
# output_file = "filled_dataset.xlsx"

# fill_blank_spaces(input_file, output_file)


from openpyxl import load_workbook, Workbook

def trim_blank_rows(input_file, output_file):
    """
    Removes blank rows from an Excel file and saves the trimmed dataset.

    :param input_file: Path to the input Excel file.
    :param output_file: Path to save the trimmed Excel file.
    """
    # Load the workbook and select the active worksheet
    wb = load_workbook(input_file)
    ws = wb.active

    # Create a new workbook for the trimmed data
    trimmed_wb = Workbook()
    trimmed_ws = trimmed_wb.active

    # Iterate through rows and keep only non-blank rows
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            trimmed_ws.append(row)

    # Save the trimmed workbook
    trimmed_wb.save(output_file)
    print(f"Trimmed dataset saved to {output_file}")

# Example usage
# input_file = "input_dataset.xlsx"
# output_file = "trimmed_dataset.xlsx"

# trim_blank_rows(input_file, output_file)

from openpyxl import load_workbook, Workbook

def filter_rows_by_values(input_file, output_file, filter_column, filter_values):
    """
    Filters rows from an Excel dataset based on values in a specified column.

    :param input_file: Path to the input Excel file.
    :param output_file: Path to save the output Excel file.
    :param filter_column: Name of the column to filter by.
    :param filter_values: List of values to retain in the filtered rows.
    """
    # Load the input workbook and select the first sheet
    wb = load_workbook(input_file)
    ws = wb.active

    # Find the index of the filter column
    header = [cell.value for cell in ws[1]]
    if filter_column not in header:
        raise ValueError(f"Column '{filter_column}' not found in the dataset.")
    filter_column_index = header.index(filter_column) + 1

    # Create a new workbook for the filtered data
    filtered_wb = Workbook()
    filtered_ws = filtered_wb.active

    # Write the header to the new sheet
    filtered_ws.append(header)

    # Filter rows based on the column value
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[filter_column_index - 1] in filter_values:
            filtered_ws.append(row)

    # Save the filtered workbook
    filtered_wb.save(output_file)
    print(f"Filtered data saved to {output_file}")

# Example usage
# input_file = "input_dataset.xlsx"
# output_file = "filtered_dataset.xlsx"
# filter_column = "City"  # Column
